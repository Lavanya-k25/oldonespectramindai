from __future__ import annotations

import json
import sys
from collections import OrderedDict
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "src" / "core" / "framework-library" / "cmmc"
SOURCE_SHEET = "Control Evidence Matrix"

SOURCE_COLUMNS = [
    "Control Family",
    "Control ID",
    "Control Requirement",
    "Evidence to Request",
    "Public Notes / Use",
    "Evidence Status",
    "Owner / Collector",
    "Date Collected",
    "Source System / Tool",
    "Notes / Gaps",
]


def cell_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def load_source_rows(source_workbook: Path) -> tuple[list[dict], list[dict], str]:
    workbook = load_workbook(source_workbook, data_only=False)
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Expected sheet '{SOURCE_SHEET}' in {source_workbook}")

    worksheet = workbook[SOURCE_SHEET]
    headers = [
        cell_value(worksheet.cell(row=1, column=column_index).value)
        for column_index in range(1, worksheet.max_column + 1)
    ]
    if headers != SOURCE_COLUMNS:
        raise ValueError(
            "Unexpected CMMC workbook columns.\n"
            f"Expected: {SOURCE_COLUMNS}\n"
            f"Actual:   {headers}"
        )

    controls = []
    for row_index in range(2, worksheet.max_row + 1):
        fields = OrderedDict(
            (
                SOURCE_COLUMNS[column_index - 1],
                cell_value(worksheet.cell(row=row_index, column=column_index).value),
            )
            for column_index in range(1, worksheet.max_column + 1)
        )
        controls.append(
            {
                "_sourceRow": row_index,
                "_sourceOrder": row_index - 1,
                "id": fields["Control ID"],
                "controlFamily": fields["Control Family"],
                "controlId": fields["Control ID"],
                "controlRequirement": fields["Control Requirement"],
                "evidenceToRequest": fields["Evidence to Request"],
                "publicNotesUse": fields["Public Notes / Use"],
                "evidenceStatus": fields["Evidence Status"],
                "ownerCollector": fields["Owner / Collector"],
                "dateCollected": fields["Date Collected"],
                "sourceSystemTool": fields["Source System / Tool"],
                "notesGaps": fields["Notes / Gaps"],
                **fields,
            }
        )

    workbook_summary = [
        {
            "sheetName": sheet.title,
            "usedRange": sheet.calculate_dimension(),
            "rowCount": sheet.max_row,
            "columnCount": sheet.max_column,
        }
        for sheet in workbook.worksheets
    ]

    return controls, workbook_summary, worksheet.calculate_dimension()


def group_by_family(records: list[dict], child_key: str) -> list[dict]:
    grouped: OrderedDict[str, list[dict]] = OrderedDict()
    for record in records:
        grouped.setdefault(record["Control Family"], []).append(record)

    return [
        {
            "controlFamily": control_family,
            "controlCount": len(items),
            child_key: items,
        }
        for control_family, items in grouped.items()
    ]


def write_json(filename: str, payload: dict) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python scripts/generate_cmmc_library.py <CMMC_Level2_Control_Evidence_Matrix.xlsx>")

    source_workbook = Path(sys.argv[1]).expanduser().resolve()
    controls, workbook_summary, used_range = load_source_rows(source_workbook)
    evidence_requirements = [
        {
            "_sourceRow": control["_sourceRow"],
            "_sourceOrder": control["_sourceOrder"],
            "id": control["Control ID"],
            "controlFamily": control["Control Family"],
            "controlId": control["Control ID"],
            "evidenceToRequest": control["Evidence to Request"],
            "publicNotesUse": control["Public Notes / Use"],
            "evidenceStatus": control["Evidence Status"],
            "ownerCollector": control["Owner / Collector"],
            "dateCollected": control["Date Collected"],
            "sourceSystemTool": control["Source System / Tool"],
            "notesGaps": control["Notes / Gaps"],
            **{column: control[column] for column in SOURCE_COLUMNS},
        }
        for control in controls
    ]
    family_summaries = [
        {
            "controlFamily": group["controlFamily"],
            "controlCount": group["controlCount"],
            "controlIds": [control["Control ID"] for control in group["controls"]],
        }
        for group in group_by_family(controls, "controls")
    ]

    source = {
        "workbookName": source_workbook.name,
        "sheetName": SOURCE_SHEET,
        "usedRange": used_range,
        "columns": SOURCE_COLUMNS,
        "workbookSummary": workbook_summary,
    }

    write_json(
        "framework.json",
        {
            "id": "cmmc-level-2",
            "name": "CMMC Level 2",
            "shortName": "CMMC",
            "version": "Level 2",
            "category": "Cybersecurity maturity",
            "description": "CMMC Level 2 framework metadata generated from the official control evidence matrix workbook.",
            "libraryStatus": "source-of-truth",
            "owner": "SpectraMind",
            "createdFor": "Framework Library",
            "source": source,
            "controlCount": len(controls),
            "controlFamilies": family_summaries,
            "notes": "Controls and evidence requirements are maintained in sibling JSON files and preserve the source workbook rows and columns.",
        },
    )
    write_json(
        "controls.json",
        {
            "source": source,
            "columns": SOURCE_COLUMNS,
            "controls": controls,
            "controlFamilies": group_by_family(controls, "controls"),
        },
    )
    write_json(
        "evidence.json",
        {
            "source": source,
            "columns": SOURCE_COLUMNS,
            "evidenceRequirements": evidence_requirements,
            "evidenceRequirementsByFamily": group_by_family(evidence_requirements, "evidenceRequirements"),
        },
    )
    write_json(
        "mappings.json",
        {
            "source": source,
            "mappings": [
                {
                    "controlId": control["Control ID"],
                    "evidenceRequirementIds": [control["Control ID"]],
                    "sourceRow": control["_sourceRow"],
                    "sourceOrder": control["_sourceOrder"],
                }
                for control in controls
            ],
        },
    )
    write_json("risks.json", {"risks": []})
    write_json("tests.json", {"tests": []})
    write_json("policies.json", {"policies": []})
    write_json("questionnaire.json", {"questionnaireSections": []})
    write_json("tasks.json", {"taskTemplates": []})
    write_json("ai-guidance.json", {"aiGuidance": []})
    write_json("audit-rules.json", {"auditRules": []})

    print(
        json.dumps(
            {
                "controls": len(controls),
                "families": len(family_summaries),
                "sourceRange": used_range,
                "columns": len(SOURCE_COLUMNS),
                "outputDir": str(OUTPUT_DIR),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
