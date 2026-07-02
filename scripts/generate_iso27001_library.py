from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = Path(r"C:\Users\komat\Downloads\ISO-27001-Controls-List-Free-Download.xlsx")
OUTPUT_DIR = ROOT / "src" / "core" / "framework-library" / "iso27001"


DOMAINS = {
    "5": {
        "annexDomain": "Organizational Controls",
        "category": "Operations",
        "prefix": "OP",
        "questionSection": "governance",
    },
    "6": {
        "annexDomain": "People Controls",
        "category": "People",
        "prefix": "PE",
        "questionSection": "people",
    },
    "7": {
        "annexDomain": "Physical Controls",
        "category": "Physical",
        "prefix": "PH",
        "questionSection": "physical",
    },
    "8": {
        "annexDomain": "Technological Controls",
        "category": "Technology",
        "prefix": "TC",
        "questionSection": "technology",
    },
}


MANDATORY_DOCUMENTS = [
    "Internal Audit",
    "Statement of Applicability",
    "ISMS Objectives",
    "Management Review",
    "Special Interest Groups",
    "Information Security Management System (ISMS) Manual",
    "Internal and External Issues and Interested Parties",
]


RISK_COUNTS = {
    "OP": 26,
    "PE": 8,
    "PH": 8,
    "TC": 25,
}


RISK_PATTERNS = {
    "OP": [
        "Incomplete information security governance can leave responsibilities unclear.",
        "Policies may be outdated or inconsistently communicated to personnel.",
        "Risk management activities may not identify or treat relevant threats.",
        "Supplier and third-party weaknesses can expose information assets.",
        "Legal, contractual, and regulatory obligations may be missed.",
        "Poor asset ownership can result in unmanaged information assets.",
        "Incident response expectations may not be coordinated or tested.",
        "Security requirements may be omitted from projects and changes.",
        "Business continuity arrangements may not support critical services.",
        "Independent reviews may not identify control weaknesses in time.",
    ],
    "PE": [
        "Personnel security responsibilities may be unclear or unenforced.",
        "Screening and onboarding gaps can grant access to unsuitable personnel.",
        "Security awareness gaps can lead to avoidable information exposure.",
        "Offboarding delays can leave access active after personnel changes.",
    ],
    "PH": [
        "Physical access weaknesses can expose facilities and equipment.",
        "Environmental threats can disrupt information processing facilities.",
        "Unprotected equipment can be damaged, lost, or used without authorization.",
        "Media handling gaps can lead to unauthorized disclosure or loss.",
    ],
    "TC": [
        "Weak logical access controls can lead to unauthorized access.",
        "Poor authentication controls can increase account compromise risk.",
        "Unmanaged vulnerabilities can leave systems exposed to exploitation.",
        "Insufficient logging and monitoring can delay detection of incidents.",
        "Weak configuration management can introduce insecure system states.",
        "Malware protection gaps can allow malicious code to spread.",
        "Inadequate backup and recovery controls can increase data loss impact.",
        "Development and change weaknesses can introduce security defects.",
    ],
}


QUESTIONNAIRE_BLUEPRINT = [
    (
        "QS-ISO-GOV",
        "ISMS Governance",
        [
            ("Q-ISO-GOV-001", "Have you defined the ISO 27001 ISMS scope?", ["Yes", "Partially", "No", "Not sure"], ["5.1", "5.4", "5.35"]),
            ("Q-ISO-GOV-002", "Do you maintain a current Statement of Applicability?", ["Yes", "Draft only", "No", "Not sure"], ["5.1", "5.36", "5.37"]),
            ("Q-ISO-GOV-003", "How often is information security risk reviewed?", ["Quarterly", "Annually", "Ad hoc", "Never"], ["5.7", "5.29", "5.35"]),
        ],
    ),
    (
        "QS-ISO-ASSET",
        "Assets and Information Handling",
        [
            ("Q-ISO-AST-001", "Do you maintain an inventory of information and associated assets?", ["Yes", "Partially", "No", "Not sure"], ["5.9", "5.10", "7.10"]),
            ("Q-ISO-AST-002", "Is information classified and labeled?", ["Yes", "Partially", "No", "Not sure"], ["5.12", "5.13"]),
            ("Q-ISO-AST-003", "Are information transfer rules documented?", ["Yes", "Partially", "No", "Not sure"], ["5.14", "8.12", "8.24"]),
        ],
    ),
    (
        "QS-ISO-PEOPLE",
        "People Security",
        [
            ("Q-ISO-PEO-001", "Are personnel screened before employment where appropriate?", ["Yes", "Partially", "No", "Not applicable"], ["6.1"]),
            ("Q-ISO-PEO-002", "Do employees complete security awareness training?", ["Yes", "Some roles only", "No", "Not sure"], ["6.3"]),
            ("Q-ISO-PEO-003", "Is access removed during transfer or termination?", ["Yes", "Partially", "No", "Not sure"], ["5.18", "6.5"]),
        ],
    ),
    (
        "QS-ISO-PHYSICAL",
        "Physical Security",
        [
            ("Q-ISO-PHY-001", "Do you control physical entry to secure areas?", ["Yes", "Partially", "No", "Not applicable"], ["7.1", "7.2", "7.3"]),
            ("Q-ISO-PHY-002", "Are equipment and media protected from loss or damage?", ["Yes", "Partially", "No", "Not sure"], ["7.4", "7.8", "7.10"]),
            ("Q-ISO-PHY-003", "Do you securely dispose of or reuse equipment?", ["Yes", "Partially", "No", "Not sure"], ["7.14", "8.10"]),
        ],
    ),
    (
        "QS-ISO-TECH",
        "Technology Security",
        [
            ("Q-ISO-TEC-001", "Is MFA required for privileged or remote access?", ["Yes", "Partially", "No", "Not sure"], ["5.17", "5.18", "8.5"]),
            ("Q-ISO-TEC-002", "Do you run vulnerability scans and track remediation?", ["Yes", "Manual only", "No", "Not sure"], ["8.8", "8.9"]),
            ("Q-ISO-TEC-003", "Are backups configured and restore-tested?", ["Yes", "Configured but not tested", "No", "Not applicable"], ["8.13", "8.14"]),
        ],
    ),
    (
        "QS-ISO-SUPPLIER",
        "Suppliers and Cloud",
        [
            ("Q-ISO-SUP-001", "Do suppliers access information assets or process company data?", ["Yes", "No", "Not sure"], ["5.19", "5.20", "5.21"]),
            ("Q-ISO-SUP-002", "Are supplier security requirements included in agreements?", ["Yes", "Partially", "No", "Not applicable"], ["5.20", "5.31", "5.32"]),
            ("Q-ISO-SUP-003", "Do you review cloud service security configurations?", ["Yes", "Partially", "No", "Not applicable"], ["5.23", "8.9", "8.20"]),
        ],
    ),
]


ISO_QUESTIONNAIRE_SECTIONS = [
    {
        "id": "QS-ISO-GENERAL",
        "title": "General Questionnaire",
        "questions": [
            {
                "id": "Q-ISO-GEN-001",
                "question": "Please provide a detailed overview of your organization",
                "type": "long_text",
                "helpText": "1. Nature of business\n2. Services and/or Products offered\n3. Customer use cases",
                "placeholder": "Describe the organization, its services or products, and customer use cases.",
                "relatedControls": ["5.1", "5.2", "5.4"],
            },
            {
                "id": "Q-ISO-GEN-002",
                "question": "Please provide a list of all the departments/teams within the organization.",
                "type": "long_text",
                "placeholder": "Example: Engineering, IT, Sales, Operations, HR, Leadership.",
                "relatedControls": ["5.2", "5.3", "5.4"],
            },
            {
                "id": "Q-ISO-GEN-003",
                "question": "ISMS Scope and Locations: Identify the locations and business units to be included in the scope of your Information Security Management System (ISMS). For each, indicate whether it is a physical site, a virtual office, or another arrangement (for example, coworking spaces).",
                "type": "long_text",
                "placeholder": "List locations, business units, and whether each is physical, virtual, or another arrangement.",
                "relatedControls": ["5.1", "5.4", "5.35"],
            },
            {
                "id": "Q-ISO-GEN-004",
                "question": "What type of work arrangement does your company use?",
                "type": "long_text",
                "helpText": "i.e. Do all employees need to be present in the office to perform their duties, are they stationed at client locations, or are they allowed to work remotely?",
                "placeholder": "Describe remote, hybrid, office, client-site, or other working arrangements.",
                "relatedControls": ["6.2", "6.3", "7.1"],
            },
            {
                "id": "Q-ISO-GEN-005",
                "question": "Are employees, contractors, partners, or anyone else who handles company information allowed to store or download any data on their devices?",
                "type": "single_select",
                "options": ["Yes", "No", "Not Applicable"],
                "relatedControls": ["5.10", "5.12", "8.1", "8.12"],
                "subQuestions": [
                    {
                        "id": "Q-ISO-GEN-005A",
                        "question": "What devices are provided by the company?",
                        "type": "long_text",
                        "placeholder": "Example: corporate laptops, phones, tablets, or no company-provided devices.",
                    }
                ],
            },
            {
                "id": "Q-ISO-GEN-006",
                "question": "Are all the devices used by employees provided by the organization, or do employees use their own personal devices for work purposes?",
                "type": "long_text",
                "placeholder": "Describe whether devices are company-owned, personal/BYOD, or a mix of both.",
                "relatedControls": ["5.10", "6.7", "8.1"],
            },
            {
                "id": "Q-ISO-GEN-007",
                "question": "Does the organization use any removable media for storing data, such as external storage devices, USB drives, or laptops? Is storage encryption enforced for such removable media?",
                "type": "long_text",
                "placeholder": "Describe removable media usage and encryption expectations.",
                "relatedControls": ["7.10", "8.10", "8.24"],
            },
        ],
    },
    {
        "id": "QS-ISO-TECHNICAL",
        "title": "Technical Questionnaire",
        "questions": [
            {
                "id": "Q-ISO-TEC-001",
                "question": "Is there a dedicated team or person responsible for handling information security in your organization?",
                "type": "single_select",
                "options": ["Yes", "No", "Not Applicable"],
                "relatedControls": ["5.2", "5.3", "5.4"],
            },
            {
                "id": "Q-ISO-TEC-002",
                "question": "Does the organization conduct any software development or testing activities?",
                "type": "single_select",
                "options": ["Yes", "No", "Not Applicable"],
                "relatedControls": ["8.25", "8.26", "8.29"],
                "subQuestions": [
                    {
                        "id": "Q-ISO-TEC-002A",
                        "question": "What type of application is developed (please specify the number of such applications)?",
                        "type": "multi_select",
                        "options": ["Web Application", "Desktop Application", "IOS App", "Android App", "Other (please specify)"],
                    }
                ],
            },
            {
                "id": "Q-ISO-TEC-003",
                "question": "Do you store, process, or transmit any internal, confidential, or sensitive information, personal identifiable of your clients or employees?",
                "type": "single_select",
                "options": ["Yes", "No", "Not Applicable"],
                "relatedControls": ["5.12", "5.13", "8.11", "8.12"],
                "subQuestions": [
                    {
                        "id": "Q-ISO-TEC-003A",
                        "question": "Please provide details on the type of data involved and location where it is stored (in house or cloud environment)",
                        "type": "long_text",
                        "placeholder": "Describe data types and storage locations.",
                    }
                ],
            },
            {
                "id": "Q-ISO-TEC-004",
                "question": "If applicable, Please provide an overview of the network architecture used within your organization. Does your organization allow the use of VPNs/Remote session software or proxy sessions?",
                "type": "long_text",
                "placeholder": "Describe internal networks, cloud access, VPNs, remote sessions, proxies, and administration paths.",
                "uploadEnabled": True,
                "relatedControls": ["8.20", "8.21", "8.22"],
            },
            {
                "id": "Q-ISO-TEC-005",
                "question": "Scope Of System",
                "type": "system_table",
                "helpText": "Please provide as exhaustive list as possible so the implementation team can estimate your needs accurately.",
                "fields": [
                    "System Name",
                    "System Usage Description",
                    "System Hosted In",
                    "Name of Third Party",
                    "Does this system store, process or transit client confidential information?",
                    "Client Service Impact by the System",
                ],
                "relatedControls": ["5.9", "5.23", "8.20"],
            },
        ],
    },
]


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def control_id(value: object) -> str:
    if isinstance(value, str):
        return clean(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return f"{value:.1f}"
    return clean(value)


def priority_for(text: str) -> str:
    high_terms = "risk access incident supplier cloud vulnerability backup authentication malware logging"
    if any(term in text.lower() for term in high_terms.split()):
        return "High"
    return "Medium"


def evidence_for(title: str) -> list[str]:
    base = title.lower()
    if "policy" in base:
        return [f"{title} document", "Approval record", "Latest review record"]
    if "access" in base or "identity" in base or "authentication" in base:
        return ["Access configuration export", "Approval records", "Periodic review evidence"]
    if "risk" in base:
        return ["Risk assessment record", "Risk treatment record", "Review approval"]
    if "supplier" in base:
        return ["Supplier register", "Supplier review record", "Contract security requirements"]
    if "backup" in base or "continuity" in base or "recovery" in base:
        return ["Backup configuration", "Restore test evidence", "Continuity plan review"]
    return [f"{title} evidence", "Procedure or standard", "Review or approval record"]


def load_controls() -> list[dict]:
    wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    ws = wb["ISO 27001 Annex A Controls List"]
    controls: list[dict] = []

    for row in ws.iter_rows(min_row=8, values_only=True):
        raw_id = row[1]
        title = clean(row[2])
        objective = clean(row[3])
        cid = control_id(raw_id)
        if not cid or "." not in cid or not title:
            continue
        domain_key = cid.split(".")[0]
        domain = DOMAINS[domain_key]
        description = objective or f"Implement and maintain {title.lower()} in line with ISO 27001:2022 Annex A."
        control = {
            "id": cid,
            "title": title,
            "category": domain["category"],
            "annexDomain": domain["annexDomain"],
            "objective": objective,
            "description": description,
            "priority": priority_for(f"{title} {description}"),
            "guidance": f"Define ownership, document the expected process, retain evidence, and review {title.lower()} at planned intervals.",
            "relatedPolicies": [],
            "relatedTests": [],
            "relatedRisks": [],
            "requiredEvidence": evidence_for(title),
            "aiGuidance": f"Confirm that {title.lower()} is documented, implemented, assigned to an owner, and supported by current evidence.",
        }
        controls.append(control)

    return controls


def controls_by_prefix(controls: list[dict]) -> dict[str, list[dict]]:
    grouped = {domain["prefix"]: [] for domain in DOMAINS.values()}
    for control in controls:
        grouped[DOMAINS[control["id"].split(".")[0]]["prefix"]].append(control)
    return grouped


def generate_risks(grouped: dict[str, list[dict]]) -> list[dict]:
    risks = []
    risks_by_prefix: dict[str, list[dict]] = {}
    for prefix, count in RISK_COUNTS.items():
        domain_controls = grouped[prefix]
        patterns = RISK_PATTERNS[prefix]
        category = domain_controls[0]["category"]
        domain_risks = []
        for index in range(count):
            start = round(index * len(domain_controls) / count)
            related = [
                domain_controls[(start + offset) % len(domain_controls)]["id"]
                for offset in range(1 + (index % 3 == 0) + (index % 5 == 0))
            ]
            title = patterns[index % len(patterns)]
            domain_risks.append(
                {
                    "id": f"R-{prefix}.{index + 1}",
                    "title": title,
                    "description": f"{title} Without effective ISO 27001 controls, the organization may fail to protect information assets or meet ISMS objectives.",
                    "category": category,
                    "severity": ["Critical", "High", "High", "Medium"][index % 4],
                    "likelihood": ["High", "Medium", "Medium", "Low"][index % 4],
                    "mitigation": "Implement the mapped Annex A controls, assign an owner, collect evidence, and review residual risk.",
                    "relatedControls": related,
                    "relatedTests": [],
                    "requiredEvidence": ["Risk assessment record", "Risk treatment decision", "Mapped control evidence"],
                }
            )
        risks_by_prefix[prefix] = domain_risks
        risks.extend(domain_risks)

    for prefix, domain_controls in grouped.items():
        covered = {
            control_id
            for risk in risks_by_prefix[prefix]
            for control_id in risk["relatedControls"]
        }
        uncovered = [control for control in domain_controls if control["id"] not in covered]
        for index, control in enumerate(uncovered):
            risk = risks_by_prefix[prefix][index % len(risks_by_prefix[prefix])]
            risk["relatedControls"] = sorted(set([*risk["relatedControls"], control["id"]]))

    return risks


def make_test_title(control: dict, variant: str) -> str:
    if variant == "evidence":
        title = control["title"].strip()
        if "policy" in title.lower() or "procedure" in title.lower():
            return f"{title} document"
        return f"Evidence of {title.lower()}"
    return f"Operating review for {control['title'].lower()}"


def generate_tests(controls: list[dict], risks: list[dict]) -> list[dict]:
    risk_ids_by_control: dict[str, list[str]] = {}
    for risk in risks:
        for control_id in risk["relatedControls"]:
            risk_ids_by_control.setdefault(control_id, []).append(risk["id"])

    counters = {domain["prefix"]: 0 for domain in DOMAINS.values()}
    tests = []
    for idx, control in enumerate(controls):
        prefix = DOMAINS[control["id"].split(".")[0]]["prefix"]
        variants = ["evidence"]
        if idx < 79:
            variants.append("operation")
        for variant in variants:
            counters[prefix] += 1
            title = make_test_title(control, variant)
            tests.append(
                {
                    "id": f"T-{prefix}.{counters[prefix]}",
                    "title": title,
                    "description": f"Verify that {control['title'].lower()} is implemented and supported by suitable evidence.",
                    "category": control["category"],
                    "frequency": "Annually" if variant == "evidence" else "Quarterly",
                    "procedure": f"Inspect the current {control['title'].lower()} evidence, confirm ownership, review date, approval, and sampled implementation where applicable.",
                    "expectedResult": "Evidence is current, approved where required, and aligned to the mapped ISO 27001 control.",
                    "relatedControls": [control["id"]],
                    "relatedRisks": risk_ids_by_control.get(control["id"], [])[:5],
                    "requiredEvidence": evidence_for(control["title"]),
                }
            )
    return tests


def controls_for_document(title: str, controls: list[dict], offset: int) -> list[str]:
    lower = title.lower()
    keyword_map = {
        "statement of applicability": ["5.1", "5.31", "5.36", "5.37"],
        "information security policy": ["5.1", "5.2", "5.4"],
        "policy": ["5.1", "5.2", "5.4"],
        "isms": ["5.1", "5.2", "5.35", "5.36"],
        "objectives": ["5.1", "5.35", "5.36"],
        "document control": ["5.33", "5.34", "5.37"],
        "internal audit": ["5.35", "5.36", "5.37"],
        "management review": ["5.35", "5.36"],
        "risk": ["5.7", "5.29", "5.35"],
        "asset": ["5.9", "5.10", "7.10"],
        "access": ["5.15", "5.16", "5.17", "5.18", "8.2", "8.3", "8.5"],
        "supplier": ["5.19", "5.20", "5.21", "5.22"],
        "incident": ["5.24", "5.25", "5.26", "5.27", "5.28"],
        "continuity": ["5.29", "5.30", "8.14"],
        "disaster": ["5.29", "5.30", "8.14"],
        "backup": ["8.13", "8.14"],
        "change": ["8.9", "8.19", "8.32"],
        "development": ["8.25", "8.26", "8.27", "8.28", "8.29", "8.31"],
        "vulnerability": ["8.8", "8.9"],
        "patch": ["8.8", "8.9"],
        "malware": ["8.7"],
        "logging": ["8.15", "8.16", "8.17"],
        "cryptography": ["8.24"],
        "classification": ["5.12", "5.13"],
        "transfer": ["5.14", "8.12", "8.24"],
        "retention": ["5.33", "8.10"],
        "physical": ["7.1", "7.2", "7.3", "7.4", "7.8"],
        "visitor": ["7.2", "7.3"],
        "training": ["6.3", "6.4"],
        "termination": ["5.18", "6.5"],
        "segregation": ["5.3", "5.15"],
        "cloud": ["5.23", "8.9", "8.20"],
        "network": ["8.20", "8.21", "8.22"],
        "endpoint": ["8.1", "8.7", "8.8"],
    }
    control_ids = {control["id"] for control in controls}
    matches: list[str] = []
    for keyword, ids in keyword_map.items():
        if keyword in lower:
            matches.extend([cid for cid in ids if cid in control_ids])
    if matches:
        return list(dict.fromkeys(matches))[:6]

    return [controls[(offset + i) % len(controls)]["id"] for i in range(2)]


def generate_documents(controls: list[dict]) -> list[dict]:
    documents = []
    for index, title in enumerate(MANDATORY_DOCUMENTS):
        related = controls_for_document(title, controls, index)
        documents.append(
            {
                "id": f"DOC-{index + 1:03d}",
                "title": title,
                "description": f"Mandatory ISO 27001 implementation document or evidence package for {title.lower()}.",
                "ownerRole": "Compliance Manager",
                "reviewFrequency": "Annually",
                "relatedControls": related,
                "templateSections": ["Purpose", "Scope", "Owner", "Process", "Evidence", "Review and Approval"],
                "aiSummary": f"Prepare and maintain {title.lower()} as part of the ISO 27001 ISMS evidence set.",
            }
        )
    return documents


def generate_evidence(tests: list[dict]) -> list[dict]:
    evidence = []
    for index, test in enumerate(tests):
        evidence.append(
            {
                "id": f"EV-ISO-{index + 1:03d}",
                "title": f"{test['title']} evidence",
                "description": f"Evidence package supporting {test['title']}.",
                "collectionMethod": "Upload document, export, screenshot, ticket, or approved record.",
                "frequency": test["frequency"],
                "requiredFor": [test["title"]],
                "acceptedFileTypes": ["pdf", "docx", "xlsx", "csv", "png", "jpg"],
                "manualOrAutomatic": "Manual",
                "relatedTests": [test["id"]],
                "relatedControls": test["relatedControls"],
            }
        )
    return evidence


def generate_tasks(controls: list[dict]) -> list[dict]:
    tasks = []
    for index, control in enumerate(controls):
        tasks.append(
            {
                "id": f"TASK-ISO-{index + 1:03d}",
                "title": f"Implement ISO control {control['id']}",
                "trigger": f"{control['title']} is not assigned, implemented, or supported by evidence.",
                "priority": control["priority"],
                "recommendedOwner": "Control Owner",
                "relatedControls": [control["id"]],
                "relatedTests": [],
            }
        )
    return tasks


def generate_questionnaire(controls: list[dict]) -> list[dict]:
    control_ids = {control["id"] for control in controls}

    def normalize_question(question: dict) -> dict:
        normalized = dict(question)
        normalized["relatedControls"] = [
            control_id for control_id in question.get("relatedControls", []) if control_id in control_ids
        ]
        normalized.setdefault("required", True)
        normalized.setdefault(
            "rules",
            [
                {
                    "condition": "answer indicates missing, incomplete, not applicable, or unclear implementation context",
                    "action": "prioritize the mapped ISO controls, tests, and evidence requirements",
                }
            ],
        )
        if "subQuestions" in normalized:
            normalized["subQuestions"] = [
                {
                    **sub_question,
                    "required": sub_question.get("required", False),
                }
                for sub_question in normalized["subQuestions"]
            ]
        return normalized

    return [
        {
            "id": section["id"],
            "title": section["title"],
            "questions": [normalize_question(question) for question in section["questions"]],
        }
        for section in ISO_QUESTIONNAIRE_SECTIONS
    ]


def generate_ai_guidance(controls: list[dict]) -> list[dict]:
    guidance = []
    for control in controls:
        title = control["title"].lower()
        guidance.append(
            {
                "controlId": control["id"],
                "purpose": f"Implement ISO 27001 control {control['id']} for {control['title']}.",
                "implementationGuidance": f"Assign an owner, document the process, implement operating activities, and retain evidence for {title}.",
                "commonMistakes": [
                    "The control exists informally but is not documented.",
                    "Evidence is stale or not tied to a defined owner.",
                ],
                "auditTips": [
                    "Provide the latest approved document or record.",
                    "Show that the control is operating for the in-scope ISMS.",
                ],
                "bestPractices": [
                    "Review evidence at least annually.",
                    "Link exceptions to risk treatment decisions.",
                ],
                "suggestedEvidence": control["requiredEvidence"],
                "suggestedNextAction": f"Upload current evidence for ISO control {control['id']}.",
            }
        )
    return guidance


def build_mappings(
    controls: list[dict],
    risks: list[dict],
    tests: list[dict],
    documents: list[dict],
    evidence: list[dict],
    tasks: list[dict],
    questionnaire: list[dict],
) -> list[dict]:
    by_control = {control["id"]: {"riskIds": [], "testIds": [], "policyIds": [], "evidenceIds": [], "taskIds": [], "questionIds": []} for control in controls}

    for risk in risks:
        for cid in risk["relatedControls"]:
            by_control[cid]["riskIds"].append(risk["id"])
    for test in tests:
        for cid in test["relatedControls"]:
            by_control[cid]["testIds"].append(test["id"])
    for document in documents:
        for cid in document["relatedControls"]:
            by_control[cid]["policyIds"].append(document["id"])
    for item in evidence:
        for cid in item["relatedControls"]:
            by_control[cid]["evidenceIds"].append(item["id"])
    for task in tasks:
        for cid in task["relatedControls"]:
            by_control[cid]["taskIds"].append(task["id"])
    for section in questionnaire:
        for question in section["questions"]:
            for cid in question["relatedControls"]:
                by_control[cid]["questionIds"].append(question["id"])

    mappings = []
    for control in controls:
        entry = by_control[control["id"]]
        mappings.append({"controlId": control["id"], **{key: sorted(set(value)) for key, value in entry.items()}})
        control["relatedRisks"] = mappings[-1]["riskIds"]
        control["relatedTests"] = mappings[-1]["testIds"]
        control["relatedPolicies"] = mappings[-1]["policyIds"]

    test_ids_by_control = {control["id"]: set() for control in controls}
    for test in tests:
        for cid in test["relatedControls"]:
            test_ids_by_control[cid].add(test["id"])

    for risk in risks:
        related_tests = set()
        for cid in risk["relatedControls"]:
            related_tests.update(test_ids_by_control.get(cid, set()))
        risk["relatedTests"] = sorted(related_tests)[:8]

    return mappings


def write_json(filename: str, payload: dict) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")


def main() -> None:
    controls = load_controls()
    grouped = controls_by_prefix(controls)
    risks = generate_risks(grouped)
    tests = generate_tests(controls, risks)
    documents = generate_documents(controls)
    evidence = generate_evidence(tests)
    tasks = generate_tasks(controls)
    questionnaire = generate_questionnaire(controls)
    ai_guidance = generate_ai_guidance(controls)
    mappings = build_mappings(controls, risks, tests, documents, evidence, tasks, questionnaire)

    write_json(
        "framework.json",
        {
            "id": "iso27001-2022",
            "name": "ISO 27001:2022",
            "shortName": "ISO 27001",
            "version": "2022",
            "category": "Information security management",
            "description": "ISO/IEC 27001:2022 framework metadata for the SpectraMind framework library.",
            "issuingBody": "ISO/IEC",
            "assessmentType": "ISMS certification readiness",
            "securityDomains": [domain["annexDomain"] for domain in DOMAINS.values()],
            "libraryStatus": "draft",
            "owner": "SpectraMind",
            "createdFor": "Framework Library",
            "notes": "Generated from the supplied ISO 27001 Annex A controls workbook. Risks, tests, policies, evidence, tasks, questionnaire content, mappings, AI guidance, and audit rules are maintained in sibling JSON files.",
        },
    )
    write_json("controls.json", {"controls": controls})
    write_json("risks.json", {"risks": risks})
    write_json("tests.json", {"tests": tests})
    write_json("policies.json", {"policies": documents})
    write_json("evidence.json", {"evidenceRequirements": evidence})
    write_json("questionnaire.json", {"questionnaireSections": questionnaire})
    write_json("tasks.json", {"taskTemplates": tasks})
    write_json("mappings.json", {"mappings": mappings})
    write_json("ai-guidance.json", {"aiGuidance": ai_guidance})
    write_json("audit-rules.json", {"_comment": "ISO 27001 audit readiness and validation rules will be defined here.", "auditRules": []})

    print(
        json.dumps(
            {
                "controls": len(controls),
                "risks": len(risks),
                "tests": len(tests),
                "mandatoryDocs": len(documents),
                "evidence": len(evidence),
                "tasks": len(tasks),
                "mappings": len(mappings),
                "questionnaireSections": len(questionnaire),
                "questionnaireQuestions": sum(len(section["questions"]) for section in questionnaire),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
